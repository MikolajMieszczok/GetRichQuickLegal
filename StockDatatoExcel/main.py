import datetime
import xlsxwriter
import yfinance as yf
import openpyxl
import xlsxwriter
from datetime import datetime, timedelta

columns = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU']
workbook = xlsxwriter.Workbook("C:/Users/Mikołaj/Desktop/Enginer/SECONDEXCEL.xlsx")
worksheet = workbook.add_worksheet()
wb = openpyxl.load_workbook("C:/Users/Mikołaj/Desktop/Enginer/FIRSTEXCEL.xlsx")
ws = wb.active
max_row = 0
for row in range(1, 30):
    try:
        print(row)
        row = str(row)
        not_filtered = yf.Ticker(str(ws["B" + row].value).split(".")[0]).history(start = str(((ws["E" + row]).value)).split(" ")[0], end = str(((ws["D" + row]).value)).split(" ")[0], interval = "1d")
        #print(not_filtered)
        if not_filtered.empty:
            continue
        filtered = not_filtered["Open"]
        for col, price in enumerate(filtered):
            worksheet.write(columns[col] + row, price)
    except Exception as e:
        print(f"Error at row {row}: {e}")
workbook.close()
wb_third = openpyxl.load_workbook("C:/Users/Mikołaj/Desktop/Enginer/FIRSTEXCEL.xlsx")
ws_third = wb_third.active
workbook = xlsxwriter.Workbook("C:/Users/Mikołaj/Desktop/Enginer/THIRDEXCEL.xlsx")
worksheet = workbook.add_worksheet()
wb_second = openpyxl.load_workbook("C:/Users/Mikołaj/Desktop/Enginer/SECONDEXCEL.xlsx")
ws_second = wb_second.active
dict_people = {}
dict_stocks = {}
wb_stocks = openpyxl.load_workbook("C:/Users/Mikołaj/Desktop/Enginer/IdPeopleStocks.xlsx")
ws_stocks = wb_stocks.active
for row in range(1, 15230):
    dict_people[str(ws_stocks["C" + str(row)].value)] = ws_stocks["A" + str(row)].value
    dict_stocks[str(ws_stocks["D" + str(row)].value)] = ws_stocks["B" + str(row)].value
for row in range(1, 30):
   if str(datetime.now().strftime("%Y-%m-%d")) != str(ws_third["D" + str(row)].value):
        max_row = 11
        break
max_row = int(max_row)
max_row = max_row - 1
for x in range(1, max_row):
    new_d_p = {v: k for k, v in dict_people.items()}
    key_p = new_d_p.get(str(ws_third["A" + str(x)].value))
    print(key_p)
    division = ws_second["E" + str(x)].value
    if key_p != None:
        worksheet.write("A" + str(x), str(key_p))
        new_d_s = {v: k for k, v in dict_stocks.items()}
        key_s = new_d_s.get(str(ws_third["B" + str(x)].value))
        worksheet.write("B" + str(x), str(key_s))
        worksheet.write("C" + str(x), str(ws_third["C" + str(x)].value))
        for y in range(len(columns)-1):
            worksheet.write(str(columns[y]) + str(x), str(ws_second[str(columns[y+1] + str(x))].value/division))
workbook.close()